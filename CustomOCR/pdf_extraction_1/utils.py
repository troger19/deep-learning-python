import re
import json
import re
import unidecode
from pyzbar.pyzbar import decode
from pdf2image import convert_from_path
import cv2

from CustomOCR.pdf_extraction_1.utils import *
import requests
from os.path import splitext
import pytesseract
import numpy as np
import csv
import xlrd
import openpyxl
import math

amount_whitelist_characters = re.compile('[^0-9,.]')
reg_total_amount2 = ['UHRADE.*(?:\s)((\d+\s)\d+[,.]\d+)', 'UHRAD[EU](.*)(e|eur|EUR|€|euro)(?:\s|$)', 'SUMA(.*)(e|eur|EUR|€|euro)(?:\s|$)',
                     'UHRAD[EU].*(?:\s)(\d\s\d+[,.]\d+)','UHRAD[EU].*(?:\s)(\d+[,.]\d+)', 'CELKOM(.*)(e|eur|EUR|€|euro)(?:\s|$)']  # 'CELKOM.*(?:\s)(\d+[,.]\d+)'
reg_skk = ['CELKOM(.*)SKK','SUMA(.*)SK','30,1260 SK (.*) SK','30,1260(.*)SK','30,126SKK\/1€(.*)SKK','30,12[86][S5]KK\/1€(.*)SKK','30[, ]?12[68][S5]KK(.*)SKK']
reg_dummy_cena = 'SUMAKHRADE((\d+)[.,](\d+))EUR'
reg_ecv = '(B(A|B|C|J|L|N|R|S|Y|T)|CA|D(K|S|T)|G(A|L)|H(C|E)|IL|K(A|I|E|K|M|N|S)|L(E|C|M|V)|M(A|I|L|T|Y)|N(I|O|M|R|Z)|P(B|D|E|O|K|N|P|T|U|V)|R(A|K|S|V)|S(A|B|C|E|I|K|L|O|N|P|V)|T(A|C|N|O|R|S|T|V)|V(K|T)|Z(A|C|H|I|M|V))([ |-]{0,1})([0-9]{3})([A-Z]{2})'
reg_ecv_dummy = '(B(A|B|C|J|L|N|R|S|Y|T)|CA|D(K|S|T)|G(A|L)|H(C|E)|IL|K(A|I|E|K|M|N|S)|L(E|C|M|V)|M(A|I|L|T|Y)|N(I|O|M|R|Z)|P(B|D|E|O|K|N|P|T|U|V)|R(A|K|S|V)|S(A|B|C|E|I|K|L|O|N|P|V)|T(A|C|N|O|R|S|T|V)|V(K|T)|Z(A|C|H|I|M|V))-([0-9]{3})([A-Z]{2})'
reg_iban = '[5S]K\d{2}\s*?\d{4}\s*?\d{4}\s*?\d{4}\s*?\d{4}\s*?\d{4}|SK\d{22}|CZ\d{22}|DE\d{20}|AT\d{18}|SK[0-9OS]{22}|SE\d{2}\s*?\d{4}\s*?\d{4}\s*?\d{4}\s*?\d{4}\s*?\d{4}'
reg_vin = [
    '(([a-h,A-H,j-n,J-N,p-z,P-Z,0-9]{9})([a-h,A-H,j-n,J-N,p,P,r-t,R-T,v-z,V-Z,0-9])([a-h,A-H,j-n,J-N,p-z,P-Z,0-9])\s*(\d{6}))',
    '(?=.*[0-9])(?=.*[A-z])[0-9A-z-]{17}']
reg_total_amount_number = ['\d+\s*\d+\s*\d+[.,]?\d+', '\d+[.,]?\d+']
reg_ico = ['IC.*(27082440)', 'ICO[: ]*(\d{8})', 'ICO\s*?\.?\s*?(\d{8})', 'ICO\s*?.?\s*?(\d{2} \d{3} \d{3})', 'ICO\s*?.?\s*?(\d{2}\s*?\d{3}\s*?\d{3})','(\d{8})\s*DIC:','(\d{8}).*DIC:']



def safe_cast(val, to_type, default=None):
    try:
        return to_type(val)
    except (ValueError, TypeError):
        return default

def safe_cast_cena(val):
    if val is None or '/' in val or '30,1260' in val:
        return 0
    val = re.sub(r'[^0-9.,]+', '', str(val))
    result = re.match('^0\d+', val)  # ak to zacina napr. 09...  ide o cislo.. 0.9 prejde
    if result:
        return 0
    val = val.strip('.,')
    if '.' in val and ',' in val:
        bodka_index = val.find('.')
        ciarka_index = val.find(',')
        if bodka_index < ciarka_index:
            val = str(val).replace('.', '')
            val = str(val).replace(',', '.')
        else:
            val = str(val).replace(',', '')
    elif ',' in val:
        val = str(val).replace(',', '.')

    casted_float = safe_cast(val, float, 0)
    if casted_float<float(10000000):
        return casted_float
    else:
        return 0


def extract_ico(unaccented_upper_text):
    for i in reg_ico:
        final_extraction = re.findall(i, str(unaccented_upper_text))
        count_ico = len(final_extraction)
        if count_ico == 1:
            ICO = final_extraction[0].replace(' ', '').replace(':', '')
        elif count_ico == 2:
            if int(final_extraction[0].replace(' ', '')) in ico_servisy.values():
                ICO = final_extraction[0]
            elif int(final_extraction[1].replace(' ', '')) in ico_servisy.values():
                ICO = final_extraction[1]
            else:
                continue
        else:
            continue
        return ICO

def extract_pdf_text(unaccented_upper_text, extracted_values,ico_servisy_p):
    global ico_servisy
    ico_servisy = ico_servisy_p
    extracted_values = extracted_values
    # SUMA
    if 'HRAD' not in extracted_values.get('cena_s_dph_word',''):
        extracted_values.update({'cena_s_dph': None})
        extracted_values.update({'cena_s_dph_conf': None})

    if extracted_values.get('cena_s_dph') is None:
        amount_str = try_multiple_regex(unaccented_upper_text, reg_total_amount2, 1, True)
        if amount_str:
            m = re.search('\d', amount_str)
            if m:
                amount_str = amount_str[m.start():]
                amount = re_replace(amount_str)
                if len(amount) > 4 and ',' in amount and '.' in amount:
                    amount = amount.replace('.', '')
                if '.' in amount:
                    amount = amount.replace('.', ',')
                extracted_values.update({'cena_s_dph': amount})

    # ECV
    if extracted_values.get('ecv') is None:
        license_plate_number = re.search(reg_ecv, unaccented_upper_text)
        if license_plate_number:
            extracted_values.update({'ecv': license_plate_number.group()})

    # IBAN
    if extracted_values.get('iban') is None:
        unaccented_upper_text = re.sub('[^0-9a-zA-Z]+', '', unaccented_upper_text)
        iban = re.search(reg_iban, unaccented_upper_text)
        if iban:
            iban = re.sub('\s+', '', str(iban.group()))
            extracted_values.update({'iban': iban})

    # VIN
    # vin = re.findall(reg_vin, unaccented_upper_text)
    if extracted_values.get('vin') is None:
        vin = try_multiple_regex(unaccented_upper_text, reg_vin)
        if vin:
            extracted_values.update({'vin': vin.replace(' ', '')})

    # ICO
    if extracted_values.get('ico') is None:
        ico = extract_ico(unaccented_upper_text)
        if ico:
            extracted_values.update({'ico': ico})
        else:
            if extracted_values.get('iban'):
                ico = ico_servisy.get(extracted_values.get('iban'))
                if ico:
                    extracted_values.update({'ico': ico})
                    extracted_values.update({'ico_conf': 999})

    return extracted_values


def re_replace(string):
    return re.sub(amount_whitelist_characters, '', string)


def load_target_values(filename):
    with open('validation\\validation.json', encoding='utf-8') as f:
        template = json.load(f)
        for i, file in enumerate(template['files']):
            if file['filename'] == filename:
                return file['values']
        else:
            return {}


def load_target_values_excel(target_values):
    ps = openpyxl.load_workbook('%s' % target_values)
    sheet = ps['Sheet1']
    total_info = {}
    for row in range(2, sheet.max_row + 1):
        filename = sheet['A' + str(row)].value
        nazov = sheet['B' + str(row)].value
        ico = sheet['C' + str(row)].value
        iban = sheet['D' + str(row)].value
        cislo_fakt = sheet['E' + str(row)].value
        datum_dodania = sheet['F' + str(row)].value
        datum_vyhotovenia = sheet['G' + str(row)].value
        ecv = sheet['H' + str(row)].value
        cez_bez_dph = sheet['I' + str(row)].value
        cena_s_dph = sheet['J' + str(row)].value
        mena = sheet['K' + str(row)].value
        vin = sheet['L' + str(row)].value

        total_info.setdefault(filename, {'nazov': None,
                                         'ico': '',
                                         'iban': '',
                                         'cislo_fakt': '',
                                         'datum_dodania': '',
                                         'datum_vyhotovenia': '',
                                         'cez_bez_dph': '',
                                         'cena_s_dph': '',
                                         'mena': '',
                                         'ecv': '',
                                         'vin': ''
                                         })

        total_info[filename]['nazov'] = nazov
        total_info[filename]['ico'] = ico
        total_info[filename]['iban'] = iban
        total_info[filename]['cislo_fakt'] = cislo_fakt
        total_info[filename]['datum_dodania'] = datum_dodania
        total_info[filename]['datum_vyhotovenia'] = datum_vyhotovenia
        total_info[filename]['cez_bez_dph'] = cez_bez_dph
        total_info[filename]['cena_s_dph'] = None if cena_s_dph is None else str(cena_s_dph).replace('.', ',')
        total_info[filename]['mena'] = mena
        total_info[filename]['ecv'] = ecv
        total_info[filename]['vin'] = vin

    return total_info





def calculate_accuracy(filename, target_values, extraction_method, extracted_values, elapsed_time):
    print('extracted_values =>', extracted_values)
    if not target_values:
        print('Nepodarilo sa nacitat cielove hodnoty pre porovnanie')
    else:
        print('target_values =>', target_values)
        count_hit = 0
        count_all = 0
        for key in target_values.keys():
            if target_values.get(key) is not None:
                count_all += 1
                if target_values.get(key) == extracted_values.get(key):
                    count_hit += 1
        accuracy = count_hit / count_all * 100
        print('presnost = ', accuracy, ' %')
    save_to_csv(filename, extraction_method, target_values, extracted_values, elapsed_time)


def extract_qr_code(full_path, extraction_method):
    extracted_values = {}
    extension = splitext(full_path)[1]
    qr = None
    if extension.lower() in ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif'):
        img = cv2.imdecode(np.fromfile(full_path, dtype=np.uint8), -1)
        qr = decode(img)
    elif extension.lower() == '.pdf':
        img = convert_from_path(full_path)
        for pageNumber, page in enumerate(img):
            qr = decode(page)
            if qr and qr[0][1] == 'QRCODE':
                break
    else:
        print('Zadany format dokumentu nie je podporovany')

    if qr and qr[0][1] == 'QRCODE':  # ak naslo zober prvy
        qrcode = qr[0].data.decode('utf-8')
        if re.match('^SPD\*\d\.\d\*', qrcode):  # QR kod typu SPD - Short Payment Descriptor
            extraction_method = 'QR - SPD'
            extract_spd(qrcode, extracted_values)
        else:
            ekasa_response = requests.post('https://ekasa.financnasprava.sk/mdu/api/v1/opd/receipt/find',
                                           json={"receiptId": qrcode})
            if ekasa_response and ekasa_response.json()['receipt']:
                cena_s_dph = str(ekasa_response.json()['receipt']['totalPrice']).replace('.', ',')
                extracted_values.update(
                    {'cena_s_dph': cena_s_dph})
                extracted_values.update({'cena_s_dph_conf': '100'})
                extracted_values.update({'cena_s_dph_word': 'UHRADA_QR|100|'+cena_s_dph})
                extraction_method = 'QR - Ekasa'
            else:
                print('PAY-By-Square')
                extraction_method = 'QR - PayBySquare'
                pay_by_square_response = requests.get('http://localhost:8102/pdf/v1/qrcode/' + qrcode)
                if pay_by_square_response and pay_by_square_response.json()['payments']:
                    cena_s_dph = str(pay_by_square_response.json()['payments'][0]['amount']).replace('.', ',')
                    extracted_values.update({'cena_s_dph': cena_s_dph})
                    extracted_values.update(
                        {'iban': pay_by_square_response.json()['payments'][0]['bankAccounts'][0]['iban']})
                    extracted_values.update({'cena_s_dph_conf': '100'})
                    extracted_values.update({'cena_s_dph_word': 'UHRADA_QR|100|'+cena_s_dph})

    return extracted_values, extraction_method


# vytiahni potrebne fieldy z SPD formatu, ktory je v tvare SPD*1.0*ACC:CZ64030*AM:6897*CC:CZK*DT:2020031....
def extract_spd(qrcode, extracted_values):
    split = qrcode.split('*')
    spd = {}
    for s in split[2:-1]:
        s_split = s.split(':')
        spd.update({s_split[0]: s_split[1]})
    cena_s_dph = spd['AM']
    if '.' in cena_s_dph:
        cena_s_dph = str(cena_s_dph).replace('.',',')
    extracted_values.update({'cena_s_dph': cena_s_dph})
    extracted_values.update({'cena_s_dph_conf': '100'})
    extracted_values.update({'cena_s_dph_word': 'UHRADA_QR|100|' + cena_s_dph})


def do_dummy_matching(dynamic_fields, extracted_dynamic_fields, extracted_dynamic_fields1, extracted_dynamic_fields2,
                      extracted_dynamic_fields4,ico_servisy, filename):
    all_text_from_all_versions = re.sub('nan', '', str(extracted_dynamic_fields['text'].values) + (
        str(extracted_dynamic_fields1['text'].values)) + (str(extracted_dynamic_fields2['text'].values)) + (str(extracted_dynamic_fields4['text'].values)))
    all_text_from_all_versions = re.sub('[^0-9a-zA-Z,.-]+', '', all_text_from_all_versions)
    for i in dynamic_fields:
        if i == 'cena_s_dph' and dynamic_fields[i] is None:
            amount_str = try_multiple_regex(all_text_from_all_versions, [reg_dummy_cena], 1, True)
            if amount_str:
                m = re.search('\d', amount_str)
                if m:
                    amount_str = amount_str[m.start():]
                    amount = re_replace(amount_str).replace('.', ',')
                    dynamic_fields.update({'cena_s_dph': amount})
        if i == 'iban' and dynamic_fields[i] is None:
            iban = re.search(reg_iban, all_text_from_all_versions)
            if iban:
                iban = iban.group()
                iban = iban.replace('O','0')
                if iban.startswith('5'):
                    iban = 'S' + iban[1:]
                dynamic_fields.update({i: iban})

        if i == 'ecv' and dynamic_fields[i] is None:
            ecv_in_filename = re.search(reg_ecv, filename)
            if ecv_in_filename:
                dynamic_fields.update({i: ecv_in_filename.group()})
            else:
                ecv_in_text = re.search(reg_ecv_dummy, all_text_from_all_versions)
                if ecv_in_text:
                    dynamic_fields.update({i: ecv_in_text.group()})

        if i == 'vin' and dynamic_fields[i] is None:
            replace = all_text_from_all_versions.replace(' ', '').replace('\n', '').strip()
            search = try_multiple_regex(replace, reg_vin)
            if search:
                dynamic_fields.update({i: search})
        if i == 'ico' and dynamic_fields[i] is None:
            if dynamic_fields.get('iban'):
                ico = ico_servisy.get(dynamic_fields.get('iban'))
                if ico:
                    dynamic_fields.update({i: ico})
    return dynamic_fields


def safe_elem(array,position):
    return None if array is None else array[position]

# dynamicke vytahovanie hodnot podla regex
def extract_dynamic_fields(image, phrases, extracted_values, ico_servisy_p, filename):
    global ico_servisy
    ico_servisy = ico_servisy_p
    dynamic_fields = extracted_values
    try:
        rotation = pytesseract.image_to_osd(image,output_type='dict')['rotate']
        if rotation !=0:
            print('rotation', rotation)
        if rotation == 90:
            image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
    except:
        print('Nastal problem so zistenim rotacie')
    if (len(image.shape) > 2):
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray_image = image

    height, width = image.shape[:2]
    # converting image into gray scale image
    resize_factor = 1

    threshold_img = cv2.threshold(gray_image, 125, 255, cv2.THRESH_BINARY)[1]
    threshold_img1 = cv2.threshold(gray_image, 100, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    blured1 = cv2.medianBlur(gray_image, 3)
    blured2 = cv2.medianBlur(gray_image, 51)
    divided = np.ma.divide(blured1, blured2).data
    normed = np.uint8(255 * divided / divided.max())
    threshed_normed = cv2.threshold(normed, 100, 255, cv2.THRESH_OTSU)[1]
    image_resize = cv2.resize(image, (int(width / resize_factor), int(height / resize_factor)))
    threshold_img_resize = cv2.resize(threshold_img, (int(width / resize_factor), int(height / resize_factor)))
    threshold_img1_resize = cv2.resize(threshold_img1, (int(width / resize_factor), int(height / resize_factor)))
    threshed_resize = cv2.resize(threshed_normed, (int(width / resize_factor), int(height / resize_factor)))
    # cv2.imshow('image ',image_resize)
    # cv2.waitKey(0)
    # cv2.imshow('threshold_img ', threshold_img_resize)
    # cv2.waitKey(0)
    # cv2.imshow('threshold_img1 ', threshold_img1_resize)
    # cv2.waitKey(0)
    # cv2.imshow('threshed ',threshed_resize)
    # cv2.waitKey(0)

    # extract fields
    extracted_dynamic_fields = pytesseract.image_to_data(image, lang='SLK', output_type='data.frame') # config='-c tessedit_char_whitelist=0123456789'
    # extracted_dynamic_fields_temp = pytesseract.image_to_data(image, lang='SLK')
    extracted_dynamic_fields_string= pytesseract.image_to_string(image, lang='SLK')
    extracted_dynamic_fields1 = pytesseract.image_to_data(threshold_img, lang='SLK', output_type='data.frame')
    # extracted_dynamic_fields1 = None
    # extracted_dynamic_fields_temp1 = pytesseract.image_to_data(threshold_img, lang='SLK')
    extracted_dynamic_fields2 = pytesseract.image_to_data(threshold_img1, lang='SLK', output_type='data.frame')
    extracted_dynamic_fields4 = pytesseract.image_to_data(threshed_normed, lang='SLK', output_type='data.frame')
    # extracted_dynamic_fields2 = None
    # extracted_dynamic_fields_temp2 = pytesseract.image_to_data(threshold_img1, lang='SLK')
    # print(extracted_dynamic_fields_temp)
    # print(extracted_dynamic_fields_temp1)
    # print(extracted_dynamic_fields_temp2)
    # add as key:value pair
    if 'HRAD' not in extracted_values.get('cena_s_dph_word',''):
        dynamic_fields.update({'cena_s_dph': None})
        dynamic_fields.update({'cena_s_dph_conf': None})
    for i, (key, value) in enumerate(phrases.items()):
        if dynamic_fields.get(key) is None:
            if key == 'cena_s_dph':
                cena1 = check_and_extract(key, value, extracted_dynamic_fields)
                cena2 = check_and_extract(key, value, extracted_dynamic_fields1)
                cena3 = check_and_extract(key, value, extracted_dynamic_fields2)
                cena4 = check_and_extract(key, value, extracted_dynamic_fields4)

                cena5 = try_multiple_regex(extracted_dynamic_fields_string,['Celková fakturovaná suma: EUR ((\d+)(\s)?(\d+)[,.](\d+))'],True)

                vysl = [safe_cast_cena(safe_elem(cena1,0)), safe_cast_cena(safe_elem(cena2,0)),
                        safe_cast_cena(safe_elem(cena3,0)), safe_cast_cena(safe_elem(cena4,0))]
                conf = [safe_elem(cena1,1), safe_elem(cena2,1), safe_elem(cena3,1), safe_elem(cena4,1)]
                words = [safe_elem(cena1,2), safe_elem(cena2,2),safe_elem(cena3,2),safe_elem(cena4,2)]
                all_same = all(x == words[0] for x in words if x is not None)
                if not all_same:
                    new_conf = []
                    new_vysl = []
                    new_words = []
                    for i, w in enumerate(words):
                        if w is not None and 'HRAD' in w:
                            new_conf.append(conf[i])
                            new_vysl.append(vysl[i])
                            new_words.append(w)

                    if new_conf:
                        conf = new_conf
                        vysl = new_vysl
                        words = new_words

                max_cena_conf =None if all(v is None for v in words) else max(i for i in conf if i is not None)
                max_cena = 0 if max_cena_conf == None else vysl[conf.index(max_cena_conf)]
                word = '' if max_cena_conf == None else words[conf.index(max_cena_conf)]

                target_word=None if max_cena==0 else [str(max_cena).replace('.',',')]
                dynamic_fields.update({'cena_s_dph_conf': max_cena_conf})
                temp = dynamic_fields.get('cena_s_dph_word','')
                dynamic_fields.update({'cena_s_dph_word': temp + word + '|' + str(max_cena_conf) + '|'
                 + ('0' if target_word is None else target_word[0]) + ';'})
            else:
                target_word = check_and_extract(key, value, extracted_dynamic_fields)
                if not target_word:
                    target_word = check_and_extract(key, value, extracted_dynamic_fields1)
                if not target_word:
                    target_word = check_and_extract(key, value, extracted_dynamic_fields2)
                if not target_word:
                    target_word = check_and_extract(key, value, extracted_dynamic_fields4)

            # if target_word is not None:
            dynamic_fields.update({key: None if target_word is None else target_word[0]})
            if key == 'iban':
                dynamic_fields.update({'iban_conf': None if target_word is None else target_word[1]})
            if key == 'ico':
                dynamic_fields.update({'ico_conf': None if target_word is None else target_word[1]})

    # v pripade ze su niektore fieldy prazdne, tak skusit este regexp na cely text
    dynamic_fields = do_dummy_matching(dynamic_fields, extracted_dynamic_fields, extracted_dynamic_fields1,
                                       extracted_dynamic_fields2,extracted_dynamic_fields4,ico_servisy,filename)
    # return JSON like object with phrase name and extracted values (amounts in EUR, etc..)  -> sum_total:20,26
    return dynamic_fields


# extrahovanie textu cez Tesseract podla klucovych slov. Extrahovanie konkretneho riadku a vyparsovanie cielovej hodnoty
def check_and_extract(key, phrase, extracted_dynamic_fields):
    if not extracted_dynamic_fields['text'].isnull().values.all():
        correct_row,conf,word = extract_correct_row(extracted_dynamic_fields, phrase)
        if correct_row:  # find final value in a row through regexp
            return find_final_value(str(correct_row), key),conf,word
    else:
        return '-',0,''


# hladanie konkretnej hodnoty podla zadefinovaneho typu. Vstupom je cely riadkok extrahovaneho textu, v ktorom by sa mala nachadzat hladana hodnota
def find_final_value(row, template_type):
    if template_type == 'cena_s_dph':
        amount = try_multiple_regex(row, reg_total_amount_number, None, True)
        if amount:
            return amount.replace('.', ',')
        else:
            return '-'
    elif template_type == 'iban':
        row1 = re.sub(r'[^a-zA-Z0-9]', '', row)
        row1 = row1.replace('O', '0')
        re_search = re.search(reg_iban, str(row1))
        if re_search:
            group = re_search.group()
            if group.startswith('5'):
                group = 'S' + group[1:]
            return group
        else:
            return None
    elif template_type == 'vin':
        replace = re.sub('\s+', '', str(row)).replace('\n', '').strip()
        search = try_multiple_regex(replace, reg_vin)
        # search = re.search(reg_vin, replace)
        if search:
            return search
        else:
            return '-'
    elif template_type == 'ecv':
        search = re.search(reg_ecv, str(row))
        if search:
            return search.group()
        else:
            return None
    elif template_type == 'ico':
        return extract_ico(row)



# hlada sa dana fraza ku ktorej sa priradi jej koordinat TOP.. nasledne sa vytiahnu vsetky slova v riadku podla predpokladu, ze ich TOP hodnota pripadne do intervalu [TOP-8, TOP+100]
def extract_correct_row(extracted_dynamic_fields, phrase):
    phrase_split = phrase.split()  # split search phrase into words
    text_split = []
    rows = []
    row_text=''
    word =''
    conf_index = 0
    for i in phrase_split:
        text_split.append(i)
        text_split.append(i + ':')
    extracted_dynamic_fields['text'] = extracted_dynamic_fields['text'].str.upper()
    for word in text_split:
        rows = []
        block_num = extracted_dynamic_fields[extracted_dynamic_fields['text'] == word]['block_num'].values
        line_num = extracted_dynamic_fields[extracted_dynamic_fields['text'] == word]['line_num'].values
        par_num = extracted_dynamic_fields[extracted_dynamic_fields['text'] == word]['par_num'].values
        word_num = extracted_dynamic_fields[extracted_dynamic_fields['text'] == word]['word_num'].values
        biggest_word_index=-1
        if block_num.size>0 and line_num.size>0:
            words_in_line = sum(list(extracted_dynamic_fields[extracted_dynamic_fields['block_num'] == block_num[-1]]['line_num'] == line_num[biggest_word_index]))
            if len(block_num)>1:
                # block_array = [block_num[max_ele_index[0]],block_num[max_ele_index[0]]+1,block_num[max_ele_index[1]],block_num[max_ele_index[1]]+1]
                block_array = [block_num[-1],block_num[-2]]
                par_array = [par_num[-1],par_num[-2]]
                line_array = [line_num[-1],line_num[-2]]
            else:
                block_array = [block_num[-1]]
                par_array = [par_num[-1]]
                line_array = [line_num[-1]]

            if 'SUMA' in phrase_split:

                search_text = extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_array))
                & (extracted_dynamic_fields['par_num'].isin(par_array))
                & (extracted_dynamic_fields['line_num'].isin(line_array))
                    # & (extracted_dynamic_fields['word_num'].isin(range(word_num[0],words_in_line)))
                                         ]['text'].values

                left = \
                    extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_array))
                                             & (extracted_dynamic_fields['par_num'].isin(par_array))
                                             & (extracted_dynamic_fields['line_num'].isin(line_array))
                        # & (extracted_dynamic_fields['word_num'].isin(range(word_num[0],words_in_line)))
                                             ]['left'].values

                confidences = \
                    extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_array))
                                             & (extracted_dynamic_fields['par_num'].isin(par_array))
                                             & (extracted_dynamic_fields['line_num'].isin(line_array))
                                         # & (extracted_dynamic_fields['word_num'].isin(range(1, 20)))
                                         ]['conf'].values
                res = {left[i]: search_text[i] for i in range(len(left))}
                res_conf = {left[i]: confidences[i] for i in range(len(left))}
                new_values=[]
                new_confidences=[]
                keyList = list(res.keys())

                skk = try_multiple_regex(str(search_text).replace("'",''), reg_skk, 1, True)

                from itertools import islice
                lit = iter(enumerate(keyList))
                for i, v in enumerate(lit):
                    if v[0] < len(keyList) - 1:
                        next_key = keyList[v[0] + 1]
                        if int(next_key) - int(v[1]) < 120 and (len(str(res[v[1]])) <= 2 and str(res[v[1]]).isdigit()):
                            new_values.append(str(res[v[1]]) + str(res[next_key]))
                            new_confidences.append(res_conf[v[1]])
                            next(islice(lit, 1, 1), None)
                        else:
                            new_values.append(res[v[1]])
                            new_confidences.append(res_conf[v[1]])
                    else:
                        new_values.append(res[v[1]])
                        new_confidences.append(res_conf[v[1]])

                h = []
                new_values = [x for x in new_values if x == x]  # odstran vsetky nan
                for i, v in enumerate(new_values):
                    new_values[i] = re.sub('30[,]?12[86][S5]KK[\/1€]?', '', v)
                for text in new_values:
                    h.append(safe_cast_cena(text))

                if skk:
                    skk = skk.replace("'",'')
                    skk_float = safe_cast_cena(skk)
                    if skk_float in h:
                        h.remove(skk_float)

                if h:
                    rows.append(max(h))
                    conf_index = np.argmax(h)
                # print(search_text)
                if len(rows)>0 and not ((len(rows)==1 and rows[0]==0)):
                    return max(rows), new_confidences[conf_index],word
            if 'ICO' in phrase_split:
                search_text = extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_num))
                            & (extracted_dynamic_fields['line_num'].isin(line_num))
                            & (extracted_dynamic_fields['par_num'].isin(par_num))
                                                       ]['text'].values

                confidences = \
                    extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_array))
                                             & (extracted_dynamic_fields['par_num'].isin(par_array))
                                             & (extracted_dynamic_fields['line_num'].isin(line_array))
                        # & (extracted_dynamic_fields['word_num'].isin(range(1, 20)))
                                             ]['conf'].values
                conf_ind = np.where(search_text==word)
                aa = []
                for i in conf_ind:
                    try:
                        aa = np.append(aa, confidences[i + 1])
                    except:
                        break
                non_zero_ico_conf = sum(map(lambda x : x > 0, aa))
                if non_zero_ico_conf==1:
                    conf_index = int(aa[np.where(aa>0)])
                elif non_zero_ico_conf > 1:
                    conf_index = 999
                else:
                    conf_index = 0

                if len(search_text)>0:
                    row_text = row_text + str(search_text)
                    # break
            if 'IBAN' in phrase_split or 'ECV' in phrase_split:
                search_text = extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_num))
                                                    & (extracted_dynamic_fields['line_num'].isin(line_num))
                                                    & (extracted_dynamic_fields['par_num'].isin(par_num))
                                                    & (extracted_dynamic_fields['word_num'].isin(
                    range(word_num[0], words_in_line)))
                                                    ]['text'].values

                confidences = \
                    extracted_dynamic_fields[(extracted_dynamic_fields['block_num'].isin(block_array))
                                             & (extracted_dynamic_fields['par_num'].isin(par_array))
                                             & (extracted_dynamic_fields['line_num'].isin(line_array))
                        # & (extracted_dynamic_fields['word_num'].isin(range(1, 20)))
                                             ]['conf'].values
                regex = re.compile("^[5S]K\d+")
                idxs = [i for i, item in enumerate(search_text) if re.search(regex, item)]
                if idxs:
                    conf_ind = idxs[0]
                    IBAN_seg_len = len(search_text[conf_ind])
                    if IBAN_seg_len >20:  # cele cislo IBANU je v 1 bunke
                        conf_index = confidences[conf_ind+1]
                    else:
                        while IBAN_seg_len<20:
                            IBAN_seg_len = IBAN_seg_len + len(search_text[conf_ind+1])
                            conf_ind = conf_ind + 1
                        from statistics import mean
                        try:
                            conf_index = mean(confidences[range(idxs[0] + 1, conf_ind + 3)])
                        except:
                            print('Nastal problem pri vypocte pravdepodobnosti pre IBAN. Nalezeny retazec nema spravnu dlzku')

                if len(search_text) > 0:
                    row_text = row_text.join(search_text)
                    break

        # clean the row by replacing non values and spaces
    clear_row_text = str(row_text).replace('nan', '')
    clear_row_text = unidecode.unidecode(clear_row_text.upper())
    correct_row = re.sub('[^0-9a-zA-Z,.: ]+', '', clear_row_text)
    return correct_row,conf_index,word


# metoda na vyskusanie hladania pre pole regexpov.. pozor, poradie je dolezite, ak sa slovo najde loop konci, preto treba zadavat regexy od najvaic specifickej k najmenej specifickej
def try_multiple_regex(row, regexp, group=None, is_number=False):
    final_extraction = None
    for i in regexp:
        final_extraction = re.search(i, str(row))
        if final_extraction:
            if group:
                final_extraction = final_extraction.group(group).replace(' ', '').replace(':', '')
            else:
                final_extraction = final_extraction.group().replace(' ', '').replace(':', '')
            if is_number:
                if re.match('\d', final_extraction):
                    break

    return final_extraction


# ulozenie extahovanych aj cielovych hodnot do vysledneho suboru kvoli analyze presnosti extrahovania
def save_to_csv(filename, extraction_method, target_values, extracted_values, elapsed_time):
    csv_columns = ['filename', 'duration', 'method', 'ico_target', 'ico_extracted', 'ico_conf','cena_s_dph_target',
                   'cena_s_dph_extracted', 'cena_s_dph_conf', 'cena_s_dph_word','iban_target', 'iban_extracted', 'iban_conf',
                   'ecv_target', 'ecv_extracted', 'vin_target', 'vin_extracted']
    dict_data = [
        {'filename': filename, 'duration': elapsed_time, 'method': extraction_method,
         'ico_target': target_values.get('ico', ' -'), 'ico_extracted': extracted_values.get('ico', ' -'),
         'ico_conf': extracted_values.get('ico_conf', ' -'),
         'cena_s_dph_target': target_values.get('cena_s_dph', ' -'),
         'cena_s_dph_extracted': extracted_values.get('cena_s_dph', ' -'),
         'cena_s_dph_conf': extracted_values.get('cena_s_dph_conf', ' -'),
         'cena_s_dph_word': extracted_values.get('cena_s_dph_word', ' -'),
         'iban_target': target_values.get('iban', ' -'), 'iban_extracted': extracted_values.get('iban', ' -'),
         'iban_conf': extracted_values.get('iban_conf', ' -'),
         'ecv_target': target_values.get('ecv', ' -'), 'ecv_extracted': extracted_values.get('ecv', ' -'),
         'vin_target': target_values.get('vin', ' -'), 'vin_extracted': extracted_values.get('vin', ' -'),
         }]
    csv_file = "data.csv"
    try:
        with open(csv_file, 'a', newline='', encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
            for data in dict_data:
                writer.writerow(data)
    except IOError:
        print("I/O error")


def load_ico_servisy():
    ps = openpyxl.load_workbook('validation\\ico_servisy.xlsx')
    sheet = ps['ico']
    print(sheet.max_row)
    ico = []
    iban = []
    for i, row in enumerate(sheet.rows):
        ico.append(row[0].value)
        iban.append(row[1].value)
    iban_ico = dict(zip(iban, ico))

    return iban_ico
